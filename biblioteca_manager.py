#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
biblioteca_manager.py
Aggiorna un file Excel (biblioteca.xlsx) con l'elenco dei libri e lo ordina automaticamente.

Ordine di ordinamento:
1) Genere/Contesto
2) Cognome autore (A->Z)
3) Nome autore
4) Anno pubblicazione (dal più vecchio al più recente)
5) Volume (solo in caso di stesso autore/genere e stesso anno)
6) Titolo

Formato input (una riga per libro), separatore consigliato: ';' oppure '|'
Cognome;Nome;Titolo;Collana;Casa editrice;Anno;Genere/Contesto

Esempio:
Omero;—;Iliade;Classici Greci;Einaudi;1990;Letteratura greca

Uso rapido:
python biblioteca_manager.py add --file biblioteca.xlsx
# poi incolla righe e termina con Ctrl+D (macOS/Linux) oppure Ctrl+Z e Invio (Windows)

Oppure da file txt:
python biblioteca_manager.py add --file biblioteca.xlsx --in libri_da_importare.txt
"""

from __future__ import annotations
import argparse
import csv
import datetime as dt
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "Genere/Contesto",
    "Cognome Autore",
    "Nome Autore",
    "Titolo",
    "Collana",
    "Casa editrice",
    "Anno pubblicazione",
    "Volume (per ordinamento)",
    "Note",
    "Inserito il",
    "Ultimo aggiornamento",
]

TABLE_NAME = "TabLibri"
SHEET_NAME = "Libri"


ITALIAN_ORDINALS = {
    "primo": 1, "prima": 1,
    "secondo": 2, "seconda": 2,
    "terzo": 3, "terza": 3,
    "quarto": 4, "quarta": 4,
    "quinto": 5, "quinta": 5,
    "sesto": 6, "sesta": 6,
    "settimo": 7, "settima": 7,
    "ottavo": 8, "ottava": 8,
    "nono": 9, "nona": 9,
    "decimo": 10, "decima": 10,
    "undicesimo": 11, "undicesima": 11,
    "dodicesimo": 12, "dodicesima": 12,
    "tredicesimo": 13, "tredicesima": 13,
    "quattordicesimo": 14, "quattordicesima": 14,
    "quindicesimo": 15, "quindicesima": 15,
    "sedicesimo": 16, "sedicesima": 16,
    "diciassettesimo": 17, "diciassettesima": 17,
    "diciottesimo": 18, "diciottesima": 18,
    "diciannovesimo": 19, "diciannovesima": 19,
    "ventesimo": 20, "ventesima": 20,
}


ROMAN_MAP = {"i": 1, "v": 5, "x": 10, "l": 50, "c": 100, "d": 500, "m": 1000}


def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))


def norm(s: Optional[str]) -> str:
    if s is None:
        return ""
    return strip_accents(str(s)).strip().casefold()


def parse_int(s: Optional[str]) -> Optional[int]:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    m = re.search(r"\d{1,4}", s)
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def roman_to_int(roman: str) -> Optional[int]:
    roman = norm(roman)
    if not roman or any(ch not in ROMAN_MAP for ch in roman):
        return None
    total = 0
    prev = 0
    for ch in reversed(roman):
        val = ROMAN_MAP[ch]
        if val < prev:
            total -= val
        else:
            total += val
            prev = val
    return total if total > 0 else None


VOLUME_RE = re.compile(
    r"\b(?:vol\.?|volume|tomo|parte)\s*"
    r"(?P<v>(?:\d+)|(?:[ivxlcdm]+)|(?:primo|secondo|terzo|quarto|quinto|sesto|settimo|ottavo|nono|decimo|"
    r"undicesimo|dodicesimo|tredicesimo|quattordicesimo|quindicesimo|sedicesimo|diciassettesimo|"
    r"diciottesimo|diciannovesimo|ventesimo))\b",
    flags=re.IGNORECASE,
)


def parse_volume(title: str) -> Optional[int]:
    if not title:
        return None
    m = VOLUME_RE.search(title)
    if not m:
        return None
    v = m.group("v")
    v_norm = norm(v)
    if v_norm.isdigit():
        return int(v_norm)
    if v_norm in ITALIAN_ORDINALS:
        return ITALIAN_ORDINALS[v_norm]
    r = roman_to_int(v_norm)
    return r


@dataclass(frozen=True)
class Book:
    genere: str
    cognome: str
    nome: str
    titolo: str
    collana: str
    editore: str
    anno: Optional[int]
    volume: Optional[int]
    note: str
    inserito_il: str
    updated_il: str

    @property
    def key(self) -> Tuple[str, str, str, str, str, Optional[int], Optional[int]]:
        """Chiave di deduplica 'ragionevole' (puoi modificarla a piacere)."""
        return (
            norm(self.genere),
            norm(self.cognome),
            norm(self.nome),
            norm(self.titolo),
            norm(self.editore),
            self.anno,
            self.volume,
        )

    @property
    def sort_key(self) -> Tuple:
        # None years/volumes should go last within the author group.
        year = self.anno if self.anno is not None else 99999
        vol = self.volume if self.volume is not None else 99999
        return (
            norm(self.genere),
            norm(self.cognome),
            norm(self.nome),
            year,
            vol,
            norm(self.titolo),
        )


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    widths = [22, 18, 16, 38, 20, 22, 18, 20, 22, 18, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    tab = Table(displayName=TABLE_NAME, ref="A1:K1")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)


def detect_delimiter(line: str) -> str:
    if ";" in line:
        return ";"
    if "|" in line:
        return "|"
    if "\t" in line:
        return "\t"
    return ","


def parse_lines(lines: Iterable[str], interactive: bool = False) -> Tuple[List[Book], List[str]]:
    books: List[Book] = []
    warnings: List[str] = []

    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for raw in lines:
        raw = raw.strip()
        if not raw:
            continue
        if raw.startswith("#"):
            continue

        delim = detect_delimiter(raw)

        if delim == ",":
            # Use csv to respect quotes
            row = next(csv.reader([raw], delimiter=",", quotechar='"', skipinitialspace=True))
        else:
            row = [c.strip() for c in raw.split(delim)]

        # Expected: 7 fields
        if len(row) < 7:
            msg = (
                f"Riga con {len(row)} campi (servono 7): {raw}\n"
                f"Formato: Cognome{delim}Nome{delim}Titolo{delim}Collana{delim}Casa editrice{delim}Anno{delim}Genere/Contesto"
            )
            if interactive:
                warnings.append(msg + "  -> Saltata (modalità interattiva: correggi e reinserisci).")
            else:
                warnings.append(msg + "  -> Saltata.")
            continue

        # If extra fields, join middle parts into Titolo (common when delimiters appear in title)
        if len(row) > 7:
            # Assume first 2 and last 4 are fixed, merge the rest into titolo
            cognome = row[0]
            nome = row[1]
            genere = row[-1]
            anno = row[-2]
            editore = row[-3]
            collana = row[-4]
            titolo = delim.join(row[2:-4]).strip()
        else:
            cognome, nome, titolo, collana, editore, anno, genere = row[:7]

        anno_i = parse_int(anno)
        vol_i = parse_volume(titolo) or parse_volume(collana)

        inserito_il = now
        updated_il = now
        books.append(
            Book(
                genere=genere.strip(),
                cognome=cognome.strip(),
                nome=nome.strip(),
                titolo=titolo.strip(),
                collana=collana.strip(),
                editore=editore.strip(),
                anno=anno_i,
                volume=vol_i,
                note="",
                inserito_il=inserito_il,
                updated_il=updated_il,
            )
        )

    return books, warnings


def load_existing(ws) -> List[Book]:
    books: List[Book] = []
    # rows start at 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        genere, cognome, nome, titolo, collana, editore, anno, volume, note, inserito_il, updated_il = (row + (None,) * 11)[:11]
        books.append(
            Book(
                genere=str(genere or "").strip(),
                cognome=str(cognome or "").strip(),
                nome=str(nome or "").strip(),
                titolo=str(titolo or "").strip(),
                collana=str(collana or "").strip(),
                editore=str(editore or "").strip(),
                anno=parse_int(anno),
                volume=parse_int(volume),
                note=str(note or "").strip(),
                inserito_il=str(inserito_il or "").strip(),
                updated_il=str(updated_il or "").strip(),
            )
        )
    return books


def write_all(ws, books: List[Book]) -> None:
    # Clear old rows (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for b in books:
        ws.append([
            b.genere,
            b.cognome,
            b.nome,
            b.titolo,
            b.collana,
            b.editore,
            b.anno if b.anno is not None else "",
            b.volume if b.volume is not None else "",
            b.note,
            b.inserito_il,
            b.updated_il,
        ])

    # Update Excel table range
    last_row = ws.max_row
    ref = f"A1:K{max(1, last_row)}"
    # Remove and re-add table to update range
    if TABLE_NAME in ws.tables:
        del ws.tables[TABLE_NAME]
    tab = Table(displayName=TABLE_NAME, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def add_books(file_path: Path, incoming: List[Book], allow_duplicates: bool) -> Tuple[int, int]:
    ensure_workbook(file_path)
    wb = load_workbook(file_path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.active
        ws.title = SHEET_NAME
    ws = wb[SHEET_NAME]

    existing = load_existing(ws)
    existing_keys = set(b.key for b in existing)

    added = 0
    skipped = 0
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    merged = list(existing)
    for b in incoming:
        b2 = Book(
            genere=b.genere,
            cognome=b.cognome,
            nome=b.nome,
            titolo=b.titolo,
            collana=b.collana,
            editore=b.editore,
            anno=b.anno,
            volume=b.volume,
            note=b.note,
            inserito_il=b.inserito_il or now,
            updated_il=now,
        )
        if (not allow_duplicates) and (b2.key in existing_keys):
            skipped += 1
            continue
        merged.append(b2)
        existing_keys.add(b2.key)
        added += 1

    merged.sort(key=lambda x: x.sort_key)
    write_all(ws, merged)
    wb.save(file_path)
    return added, skipped


def cmd_add(args: argparse.Namespace) -> int:
    file_path = Path(args.file).expanduser().resolve()
    if args.input:
        in_path = Path(args.input).expanduser().resolve()
        lines = in_path.read_text(encoding="utf-8").splitlines()
    else:
        # Read from stdin
        lines = sys.stdin.read().splitlines()

    books, warnings = parse_lines(lines, interactive=False)
    for w in warnings:
        print("AVVISO:", w, file=sys.stderr)

    added, skipped = add_books(file_path, books, allow_duplicates=args.allow_duplicates)
    print(f"Operazione completata. Aggiunti: {added}. Duplicati saltati: {skipped}. File: {file_path}")
    return 0


def cmd_init(args: argparse.Namespace) -> int:
    file_path = Path(args.file).expanduser().resolve()
    ensure_workbook(file_path)
    print(f"Creato (o già esistente): {file_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="biblioteca_manager.py", description="Gestione catalogo libri in Excel.")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="Crea il file Excel se non esiste.")
    p_init.add_argument("--file", default="biblioteca.xlsx", help="Percorso del file Excel.")
    p_init.set_defaults(func=cmd_init)

    p_add = sub.add_parser("add", help="Aggiunge libri al file Excel e riordina.")
    p_add.add_argument("--file", default="biblioteca.xlsx", help="Percorso del file Excel.")
    p_add.add_argument("--in", dest="input", default=None, help="File di testo (UTF-8) da importare, una riga per libro.")
    p_add.add_argument("--allow-duplicates", action="store_true", help="Non saltare duplicati (li aggiunge comunque).")
    p_add.set_defaults(func=cmd_add)

    return p


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())


#By Riva Thomas.
